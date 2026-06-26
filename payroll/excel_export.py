from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from xml.sax.saxutils import escape

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center")
RIGHT       = Alignment(horizontal="right",  vertical="center")
 
MONEY_FORMAT = '#,##0.00'
DATE_FORMAT  = 'YYYY-MM-DD'
DATETIME_FMT = 'YYYY-MM-DD HH:MM'

MONEY_COLS   = {10, 11, 12, 13, 14, 15, 16, 17}
NUMERIC_COLS = {6, 7, 8, 9}    

def column_name(index):
    name = ''
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def cell_xml(row_index, column_index, value):
    reference = f'{column_name(column_index)}{row_index}'

    if isinstance(value, Decimal):
        return f'<c r="{reference}"><v>{value}</v></c>'
    if isinstance(value, (int, float)):
        return f'<c r="{reference}"><v>{value}</v></c>'
    if isinstance(value, datetime):
        value = value.strftime('%Y-%m-%d %H:%M')
    elif isinstance(value, date):
        value = value.strftime('%Y-%m-%d')
    elif value is None:
        value = ''

    return f'<c r="{reference}" t="inlineStr"><is><t>{escape(str(value))}</t></is></c>'


def worksheet_xml(rows):
    sheet_rows = []
    for row_index, row in enumerate(rows, start=1):
        cells = ''.join(cell_xml(row_index, column_index, value) for column_index, value in enumerate(row, start=1))
        sheet_rows.append(f'<row r="{row_index}">{cells}</row>')

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetData>'
        f'{"".join(sheet_rows)}'
        '</sheetData>'
        '</worksheet>'
    )


def build_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"

    for row_idx, row in enumerate(rows, start=1):
        is_header = row_idx == 1

        for col_idx, value in enumerate(row, start=1):
            # ── Coerce value types ────────────────────────────────────────
            if isinstance(value, Decimal):
                value = float(value)
 
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
 
            # ── Header row ────────────────────────────────────────────────
            if is_header:
                cell.font      = HEADER_FONT
                cell.fill      = HEADER_FILL
                cell.alignment = CENTER
                ws.row_dimensions[row_idx].height = 32
                continue
 
            # ── Data rows ─────────────────────────────────────────────────
            cell.font = BODY_FONT
 
            if isinstance(value, float) and col_idx in MONEY_COLS:
                cell.number_format = MONEY_FORMAT
                cell.alignment     = RIGHT
 
            elif isinstance(value, float) and col_idx in NUMERIC_COLS:
                cell.number_format = '#,##0.00'
                cell.alignment     = RIGHT
 
            elif isinstance(value, datetime):
                cell.number_format = DATETIME_FMT
                cell.alignment     = CENTER
 
            elif isinstance(value, date):
                cell.number_format = DATE_FORMAT
                cell.alignment     = CENTER
 
            elif isinstance(value, str):
                cell.alignment = LEFT
 
            else:
                cell.alignment = CENTER
 
        # Zebra striping for readability
        if not is_header and row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="F2F7FC")
            for col_idx in range(1, len(row) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
 
    # ── Column widths (based on content + header) ─────────────────────────
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            try:
                cell_len = len(str(cell.value or ''))
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        # Monetary columns need a bit more room for formatting
        if col_idx in MONEY_COLS:
            width = max(max_len + 6, 14)
        else:
            width = max(max_len + 3, 10)
        ws.column_dimensions[col_letter].width = min(width, 35)
 
    # ── Freeze header, add auto-filter ────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
 
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
 
